import json
import re
import sys
from copy import copy
from pathlib import Path
from zipfile import BadZipFile

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string


def load_config(path):
    config_path = Path(path).resolve()
    with config_path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    config["_base_dir"] = str(config_path.parent)
    return config


def col_index(columns, key, default):
    value = columns.get(key, default)
    if isinstance(value, int):
        return value
    return column_index_from_string(value)


def copy_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(target, MergedCell):
            continue
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)


def clear_row(ws, row):
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = None


def safe_file_name(name):
    name = re.sub(r'[<>:"/\\|?*]', "_", name).strip()
    return name or "file"


def rename_files(base_dir, folder, renames):
    folder_path = base_dir / folder
    results = []
    for old_name, new_name in renames.items():
        old_path = folder_path / old_name
        new_path = folder_path / safe_file_name(new_name)
        if old_path == new_path:
            continue
        if new_path.exists():
            results.append({"from": old_name, "to": new_path.name, "status": "exists"})
            continue
        if old_path.exists():
            old_path.rename(new_path)
            results.append({"from": old_name, "to": new_path.name, "status": "renamed"})
        else:
            results.append({"from": old_name, "to": new_path.name, "status": "missing"})
    return results


def unique_path(path):
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 2
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def rename_order_images(base_dir, items):
    results = []
    for item in items:
        order_image = item.get("order_image")
        reason = item.get("reason")
        if not order_image or not reason:
            continue

        old_path = base_dir / order_image
        if not old_path.exists():
            results.append({"from": order_image, "to": "", "status": "missing"})
            continue

        new_path = old_path.with_name(safe_file_name(reason) + old_path.suffix.lower())
        if old_path.resolve() == new_path.resolve():
            results.append({"from": order_image, "to": str(new_path.relative_to(base_dir)), "status": "already_named"})
            continue

        final_path = unique_path(new_path)
        old_path.rename(final_path)
        new_relative = str(final_path.relative_to(base_dir)).replace("\\", "/")
        item["order_image"] = new_relative
        results.append({"from": order_image, "to": new_relative, "status": "renamed"})
    return results


def create_standard_workbook(config):
    wb = Workbook()
    ws = wb.active
    ws.title = config.get("sheet", "报销单")

    widths = {
        "A": 4,
        "B": 12,
        "C": 28,
        "D": 24,
        "E": 12,
        "F": 14,
        "G": 18,
        "H": 28,
    }
    widths.update(config.get("widths", {}))
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.merge_cells("B1:H1")
    ws["B1"] = "费用报销单"
    ws["B1"].font = Font(size=18, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["F2"] = "填报日期："
    ws["G2"] = config.get("date", "")
    ws["B4"] = "报销人："
    ws["D4"] = "所属部门："
    ws["F4"] = "联系电话："

    headers = ["费用类别", "事由", "截图", "金额", "是否增值税发票", "票号", "备注"]
    header_row = int(config.get("start_row", 9)) - 1
    for offset, title in enumerate(headers, start=2):
        cell = ws.cell(header_row, offset)
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(header_row, header_row + max(len(config.get("items", [])), 1) + 2):
        for col in range(2, 9):
            ws.cell(row, col).border = border

    return wb, ws


def open_template_or_create(config):
    base_dir = Path(config["_base_dir"])
    template_name = config.get("template", "费用报销单模板.xlsx")
    skill_dir = Path(__file__).resolve().parents[1]
    template_candidates = []
    if template_name:
        configured = Path(template_name)
        template_candidates.append(configured if configured.is_absolute() else base_dir / configured)
    template_candidates.append(skill_dir / "assets" / "费用报销单模板.xlsx")

    template = next((candidate for candidate in template_candidates if candidate.exists()), None)
    if template and template.exists():
        try:
            wb = openpyxl.load_workbook(template)
            sheet_name = config.get("sheet", "Sheet1")
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            config["_template_path"] = str(template)
            return wb, ws, True
        except (BadZipFile, OSError, KeyError) as exc:
            if config.get("template_required"):
                raise
            config["_template_warning"] = f"Template ignored because it could not be opened: {exc}"
    if config.get("template_required"):
        searched = ", ".join(str(candidate) for candidate in template_candidates)
        raise FileNotFoundError(f"Template not found. Searched: {searched}")
    return (*create_standard_workbook(config), False)


def build_workbook(config):
    base_dir = Path(config["_base_dir"])
    output = base_dir / config.get("output", "报销单_整理完成.xlsx")
    items = config["items"]
    columns = config.get("columns", {})

    category_col = col_index(columns, "category", "B")
    reason_col = col_index(columns, "reason", "C")
    screenshot_col_letter = columns.get("screenshot", "D")
    amount_col = col_index(columns, "amount", "E")
    invoice_col = col_index(columns, "has_invoice", "F")
    invoice_no_col = col_index(columns, "invoice_no", "G")
    remark_col = col_index(columns, "remark", "H")
    total_col = col_index(columns, "total", "G")

    start_row = int(config.get("start_row", 9))
    original_end_row = int(config.get("original_end_row", 34))
    total_row = int(config.get("total_row", 35))

    wb, ws, used_template = open_template_or_create(config)

    if used_template:
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row >= start_row:
                ws.unmerge_cells(str(merged_range))

        available_rows = original_end_row - start_row + 1
        extra_rows = max(0, len(items) - available_rows)
        if extra_rows:
            ws.insert_rows(total_row, extra_rows)
            for row in range(total_row, total_row + extra_rows):
                copy_style(ws, original_end_row, row)

        new_total_row = total_row + extra_rows
        for row in range(start_row, new_total_row):
            copy_style(ws, start_row, row)
            clear_row(ws, row)
    else:
        new_total_row = start_row + len(items)

    end_row = start_row + len(items) - 1

    for column, width in config.get("widths", {}).items():
        ws.column_dimensions[column].width = width

    for row in range(start_row, new_total_row):
        ws.row_dimensions[row].height = config.get("image_row_height", 122) if row <= end_row else 17

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    red_font = Font(color="FF0000")
    red_fill = PatternFill("solid", fgColor=config.get("supplemental_fill", "FFF2F2"))
    yellow_fill = PatternFill("solid", fgColor=config.get("invoice_fill", "FFF2CC"))

    image_width = int(config.get("image_width", 125))
    image_height = int(config.get("image_height", 190))

    for row, item in enumerate(items, start=start_row):
        ws.cell(row, category_col).value = item.get("category", "")
        ws.cell(row, reason_col).value = item["reason"]
        ws.cell(row, amount_col).value = item.get("amount")
        ws.cell(row, invoice_col).value = "是" if item.get("has_invoice") else "否"
        if item.get("invoice_no"):
            ws.cell(row, invoice_no_col).value = item["invoice_no"]
        ws.cell(row, remark_col).value = item.get("remark", "")

        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.alignment = left if col in (reason_col, remark_col) else center
            if item.get("has_invoice"):
                cell.fill = yellow_fill
            if item.get("supplemental"):
                cell.font = red_font
                if not item.get("has_invoice"):
                    cell.fill = red_fill

        ws.cell(row, amount_col).number_format = config.get("amount_format", '¥#,##0.00')

        order_image = item.get("order_image")
        if order_image:
            image_path = base_dir / order_image
            if image_path.exists():
                image = XLImage(str(image_path))
                image.width = image_width
                image.height = image_height
                ws.add_image(image, f"{screenshot_col_letter}{row}")

    clear_row(ws, new_total_row)
    ws.cell(new_total_row, category_col).value = config.get("total_label", "金额合计")
    ws.cell(new_total_row, total_col).value = f"=SUM({columns.get('amount', 'E')}{start_row}:{columns.get('amount', 'E')}{end_row})"
    ws.cell(new_total_row, total_col).number_format = config.get("amount_format", '¥#,##0.00')
    currency_col = config.get("currency_col")
    if currency_col:
        ws.cell(new_total_row, col_index(columns, "currency", currency_col)).value = config.get("currency_text", "元")

    if config.get("date_cell") and config.get("date"):
        ws[config["date_cell"]] = config["date"]
    elif config.get("date"):
        ws["G2"] = config["date"]

    wb.save(output)
    return output, used_template


def verify_workbook(path, config):
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_name = config.get("sheet", "Sheet1")
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    start_row = int(config.get("start_row", 9))
    category_col = col_index(config.get("columns", {}), "category", "B")
    amount_col = col_index(config.get("columns", {}), "amount", "E")
    total_col = col_index(config.get("columns", {}), "total", "G")

    rows = []
    row = start_row
    while row <= ws.max_row:
        value = ws.cell(row, category_col).value
        if value == "金额合计":
            break
        if value is not None:
            rows.append(row)
        row += 1

    return {
        "rows": len(rows),
        "images": len(ws._images),
        "sum": round(sum(ws.cell(r, amount_col).value or 0 for r in rows), 2),
        "total_formula": ws.cell(row, total_col).value,
    }


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python build_reimbursement.py config.json")

    config = load_config(sys.argv[1])
    base_dir = Path(config["_base_dir"])
    order_results = []
    if config.get("rename_order_images", True):
        order_results = rename_order_images(base_dir, config.get("items", []))
    invoice_results = rename_files(
        base_dir,
        config.get("invoice_dir", "发票"),
        config.get("invoice_renames", {}),
    )
    output, used_template = build_workbook(config)
    verification = verify_workbook(output, config)
    print(json.dumps({
        "output": str(output),
        "used_template": used_template,
        "template_path": config.get("_template_path"),
        "template_warning": config.get("_template_warning"),
        "order_image_renames": order_results,
        "invoice_renames": invoice_results,
        "verification": verification,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
